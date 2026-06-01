"""Exportação do relatório consolidado em Excel (.xlsx).

Recebe o dict que `relatorios_service.dados_relatorio()` já entrega e
constrói um workbook com 3 abas: Resumo (KPIs), Peças (ranking
detalhado) e Pagamentos (distribuição por forma).

Quem chama:
    from backend.services.relatorios_export import gerar_xlsx
    buffer = gerar_xlsx(dados_relatorio(mes='2026-05'))
    return send_file(buffer, ..., download_name='relatorio.xlsx')

O `buffer` é um `BytesIO` em memória — nada é gravado em disco.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

_MOEDA = 'R$ #,##0.00'
_PORCENTAGEM = '0.0"%"'
_NEGRITO = Font(bold=True)


def gerar_xlsx(dados):
    wb = Workbook()
    _aba_resumo(wb.active, dados)
    _aba_pecas(wb.create_sheet("Peças"), dados.get("pecas_detalhe", []))
    _aba_pagamentos(wb.create_sheet("Pagamentos"), dados.get("formas_pagamento", []))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------- Abas ----------

def _aba_resumo(aba, dados):
    aba.title = "Resumo"
    aba["A1"] = "Período"
    aba["B1"] = dados.get("periodo_label", "")
    aba["A1"].font = _NEGRITO

    linhas = [
        ("Total de vendas", dados.get("total_vendas", 0), None),
        ("Faturamento", dados.get("faturamento", 0), _MOEDA),
        ("Custo de produção das vendas", dados.get("custo_total", 0), _MOEDA),
        ("Lucro bruto", dados.get("lucro_bruto", 0), _MOEDA),
        ("Despesas", dados.get("total_despesas", 0), _MOEDA),
        ("Gasto em materiais", dados.get("gasto_materiais", 0), _MOEDA),
        ("Lucro líquido", dados.get("lucro_liquido", 0), _MOEDA),
    ]
    for i, (rotulo, valor, fmt) in enumerate(linhas, start=3):
        aba.cell(row=i, column=1, value=rotulo)
        celula = aba.cell(row=i, column=2, value=valor)
        if fmt:
            celula.number_format = fmt

    aba.column_dimensions["A"].width = 32
    aba.column_dimensions["B"].width = 18


def _aba_pecas(aba, pecas):
    cabecalho = [
        "Peça", "Vendas", "Quantidade",
        "Faturamento", "Custo", "Lucro", "Margem (%)",
    ]
    _cabecalho(aba, cabecalho)

    for i, p in enumerate(pecas, start=2):
        aba.cell(row=i, column=1, value=p["peca_nome"])
        aba.cell(row=i, column=2, value=p["vendas"])
        aba.cell(row=i, column=3, value=p["quantidade"])
        aba.cell(row=i, column=4, value=p["faturamento"]).number_format = _MOEDA
        aba.cell(row=i, column=5, value=p["custo"]).number_format = _MOEDA
        aba.cell(row=i, column=6, value=p["lucro"]).number_format = _MOEDA
        aba.cell(row=i, column=7, value=p["margem"]).number_format = _PORCENTAGEM

    aba.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        aba.column_dimensions[col].width = 14


def _aba_pagamentos(aba, formas):
    _cabecalho(aba, ["Forma de pagamento", "Valor", "% do faturamento"])

    for i, f in enumerate(formas, start=2):
        aba.cell(row=i, column=1, value=f["forma"])
        aba.cell(row=i, column=2, value=f["valor"]).number_format = _MOEDA
        aba.cell(row=i, column=3, value=f["pct"]).number_format = _PORCENTAGEM

    aba.column_dimensions["A"].width = 22
    aba.column_dimensions["B"].width = 14
    aba.column_dimensions["C"].width = 18


# ---------- Helpers ----------

def _cabecalho(aba, titulos):
    for col, titulo in enumerate(titulos, start=1):
        celula = aba.cell(row=1, column=col, value=titulo)
        celula.font = _NEGRITO
        celula.alignment = Alignment(horizontal="left")
